# file_processing.py
import os
import logging
import re
import tkinter as tk # Necesare pentru create_progress_window
from tkinter import ttk # Necesare pentru create_progress_window
# Nu este nevoie de messagebox aici, este folosit în main app
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import threading
from queue import Queue, Empty # Queue este folosit, Empty nu neapărat direct de utilizator
from datetime import datetime
import pymysql
from sqlalchemy import create_engine
import io


# Expresii regulate
RE_IBAN_EXTRACT = re.compile(r"([A-Z]{2}[0-9]{2}[A-Z0-9]{11,30})")
RE_CIF = re.compile(r"C\.I\.F\.?:\s?(\d+)")
RE_FACTURA = re.compile(r"(?:FACT(?:URA)?(?: NR)?(?:\.|:)?\s*|F\.\s*)(\w+)")
RE_BENEFICIAR = re.compile(r"\b([A-Z][A-Z\s.\-0-9&]{5,})\b")
RE_TID = re.compile(r"TID:?\s?(\S+)")
RE_RRN = re.compile(r"RRN:?\s?(\S+)")
RE_PAN = re.compile(r"PAN:?\s?(\S+)")
RE_MID = re.compile(r"MID\s*(\d+)")  # Merchant ID pentru tranzacții POS grupate (BT feb 2026)

def extract_iban_from_mt940(file_path):
    """
    Extrage IBAN-ul din câmpul :25: dintr-un fișier MT940.
    Funcția este acum mai robustă și citește linie cu linie.
    Returnează IBAN-ul ca string (litere mari) sau None.
    """
    logging.debug(f"DEBUG_EXTRACT_IBAN: Se procesează fișierul (versiune îmbunătățită): {file_path}")
    iban_candidate_line_content = None
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            for line_num, line_text in enumerate(f):
                stripped_line = line_text.strip()
                if stripped_line.startswith(":25:"):
                    iban_candidate_line_content = stripped_line[4:].strip() # Preluăm conținutul de după :25:
                    logging.debug(f"DEBUG_EXTRACT_IBAN: Linia :25: găsită (linia {line_num + 1}): '{iban_candidate_line_content}'")
                    break # Am găsit primul tag :25:, ne oprim
            else: # Se execută dacă bucla for s-a terminat fără 'break'
                logging.debug(f"DEBUG_EXTRACT_IBAN: Tag-ul :25: nu a fost găsit în fișierul {os.path.basename(file_path)}")
                return None

        if iban_candidate_line_content:
            potential_iban_part = iban_candidate_line_content
            # Eliminăm orice prefix de tipul "COD BANCA/" sau similar, dacă există
            if '/' in potential_iban_part:
                potential_iban_part = potential_iban_part.split('/')[-1].strip()
                logging.debug(f"DEBUG_EXTRACT_IBAN: Parte după '/' selectată: '{potential_iban_part}'")
            
            # Eliminăm orice caractere non-alfanumerice, cu excepția literelor și cifrelor
            # (IBAN-urile standard conțin doar litere și cifre)
            cleaned_potential_iban = re.sub(r'[^A-Z0-9]', '', potential_iban_part.upper())
            logging.debug(f"DEBUG_EXTRACT_IBAN: IBAN curățat și uppercase: '{cleaned_potential_iban}'")

            # Aplicăm regex-ul pentru a valida și extrage formatul IBAN
            iban_match = RE_IBAN_EXTRACT.fullmatch(cleaned_potential_iban) # Folosim fullmatch pentru a valida întregul string curățat
            if iban_match:
                extracted_iban = iban_match.group(1) # grupul 1 este întregul IBAN potrivit
                logging.debug(f"DEBUG_EXTRACT_IBAN: IBAN extras și validat prin fullmatch: {extracted_iban}")
                return extracted_iban
            else:
                logging.debug(f"DEBUG_EXTRACT_IBAN: IBAN-ul curățat '{cleaned_potential_iban}' nu corespunde formatului așteptat de RE_IBAN_EXTRACT.")
                # Ca o ultimă încercare, dacă regex-ul e prea strict, dar stringul arată a IBAN
                if 15 <= len(cleaned_potential_iban) <= 34 and cleaned_potential_iban[:2].isalpha() and cleaned_potential_iban[2:4].isdigit():
                    logging.debug(f"DEBUG_EXTRACT_IBAN: IBAN-ul curățat trece validarea de bază. Se returnează: {cleaned_potential_iban}")
                    return cleaned_potential_iban

            logging.debug(f"DEBUG_EXTRACT_IBAN: Nu s-a găsit un IBAN valid în conținutul liniei :25: ('{iban_candidate_line_content}')")
            return None
        else:
            # Acest caz este acoperit de for-else de mai sus, dar lăsăm pentru claritate
            logging.debug(f"DEBUG_EXTRACT_IBAN: Conținutul liniei :25: este gol după extragere (improbabil dacă tag-ul a fost găsit).")
            return None

    except Exception as e:
        logging.debug(f"DEBUG_EXTRACT_IBAN: Eroare la extragerea IBAN-ului din {os.path.basename(file_path)}: {e}")
        return None

def create_progress_window(master_ref, title, message):
    progress_win = tk.Toplevel(master_ref)
    progress_win.title(title)
    progress_win.transient(master_ref)
    progress_win.grab_set()
    progress_win.geometry("400x120")
    progress_win.resizable(False, False)

    ttk.Label(progress_win, text=message, wraplength=380).pack(pady=10)
    progress_status_label_widget = ttk.Label(progress_win, text="Inițializare...")
    progress_status_label_widget.pack(pady=(0, 5), fill=tk.X, padx=10)
    progress_bar_widget = ttk.Progressbar(progress_win, orient='horizontal', length=380, mode='determinate')
    progress_bar_widget.pack(pady=5)
    
    master_x = master_ref.winfo_x()
    master_y = master_ref.winfo_y()
    master_w = master_ref.winfo_width()
    master_h = master_ref.winfo_height()
    win_w, win_h = 400, 120
    x = master_x + (master_w - win_w) // 2
    y = master_y + (master_h - win_h) // 2
    progress_win.geometry(f"+{x}+{y}")

    return progress_win, progress_bar_widget, progress_status_label_widget

def threaded_import_worker(app_instance, file_paths, q_ref, active_account_id_for_import, db_credentials):
    logging.debug(f"DEBUG_THREAD: Pornit threaded_import_worker. Cont țintă ID: {active_account_id_for_import}")
    inserted, ignored = 0, 0
    thread_conn_local = None
    try:
        # === AICI ESTE SINGURA MODIFICARE LOGICĂ ===
        # Am eliminat dependența de `app_instance` și folosim direct `db_credentials`.
        if not db_credentials:
            raise ConnectionError("Credentialele DB nu au fost furnizate worker-ului de import.")

        if active_account_id_for_import is None:
            raise ValueError("ID-ul contului activ nu a fost furnizat pentru import.")

        # Construim parametrii de conexiune din dicționarul primit
        conn_params = db_credentials.copy()
        conn_params['db'] = conn_params.pop('database', None)
        conn_params['passwd'] = conn_params.pop('password', None)
        conn_params['charset'] = 'utf8mb4'
        
        # Ne conectăm folosind parametrii locali
        thread_conn_local = pymysql.connect(**conn_params)
        cursor = thread_conn_local.cursor()
        # === SFÂRȘITUL MODIFICĂRII LOGICE. RESTUL CODULUI ESTE IDENTIC CU ORIGINALUL. ===

        for i, file_path in enumerate(file_paths):
            q_ref.put(("progress", i, f"Procesare: {os.path.basename(file_path)}"))
            
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()

            cursor.execute("SELECT cod FROM tipuri_tranzactii")
            known_tx_types = {row[0] for row in cursor.fetchall()}
            
            potential_new_tx_codes = set(re.findall(r":61:\d+[CD][\d,]+([A-Z]{4})", content))
            new_codes_to_add = potential_new_tx_codes - known_tx_types
            
            if new_codes_to_add:
                cursor.execute("SELECT cod_swift, descriere_standard FROM swift_code_descriptions")
                swift_descriptions = {row[0]: row[1] for row in cursor.fetchall()}

                for new_code in new_codes_to_add:
                    description = swift_descriptions.get(new_code, f"Tip nou, cod: {new_code}")
                    cursor.execute("INSERT INTO tipuri_tranzactii (cod, descriere_tip) VALUES (%s, %s)", (new_code, description))
                
                thread_conn_local.commit()

            transactions = re.findall(r"(:61:.*?)(?=(:61:|$))", content, re.DOTALL)
            
            for tx_block, _ in transactions:
                match61 = re.search(r":61:(\d{6})(?:\d{4})?([CD])([\d,]+)([A-Z]{4})", tx_block)
                if not match61: continue

                date_str, type_char, amount_str, tx_code_full = match61.groups()
                amount = float(amount_str.replace(',', '.'))
                tx_type = "credit" if type_char == 'C' else "debit"
                date_obj = datetime.strptime(date_str, '%y%m%d').date()
                
                match86 = re.search(r":86:(.*?)$", tx_block, re.DOTALL)
                full_descr = match86.group(1).strip().replace('\n', ' ') if match86 else ""

                cif_match = RE_CIF.search(full_descr)
                beneficiar_match = RE_BENEFICIAR.search(full_descr)
                factura_match = RE_FACTURA.search(full_descr)
                tid_match = RE_TID.search(full_descr)
                rrn_match = RE_RRN.search(full_descr)
                pan_match = RE_PAN.search(full_descr)
                mid_match = RE_MID.search(full_descr)

                cif_val = cif_match.group(1).strip() if cif_match else None
                beneficiar_val = beneficiar_match.group(1).strip() if beneficiar_match else None
                factura_val = factura_match.group(1).strip() if factura_match else None
                tid_val = tid_match.group(1).strip() if tid_match else None
                rrn_val = rrn_match.group(1).strip() if rrn_match else None
                pan_val = pan_match.group(1).strip() if pan_match else None
                mid_val = mid_match.group(1).strip() if mid_match else None

                cursor.execute("SELECT 1 FROM tranzactii WHERE data=%s AND suma=%s AND tip=%s AND descriere=%s AND id_cont_fk=%s",
                               (date_obj.strftime('%Y-%m-%d'), amount, tx_type, full_descr, active_account_id_for_import))
                if not cursor.fetchone():
                    sql_insert = ("INSERT INTO tranzactii (id_cont_fk, data, descriere, suma, tip, cod_tranzactie_fk, cif, beneficiar, factura, tid, rrn, pan, mid) "
                                  "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                    values = (active_account_id_for_import, date_obj.strftime('%Y-%m-%d'), full_descr, amount,
                              tx_type, tx_code_full, cif_val, beneficiar_val, factura_val, tid_val, rrn_val, pan_val, mid_val)
                    cursor.execute(sql_insert, values)
                    inserted += 1
                else:
                    ignored += 1

        if inserted > 0:
            thread_conn_local.commit()
        
        cursor.close()
        q_ref.put(("done", "import_batch", (inserted, ignored)))

    except pymysql.Error as e:
        error_message = f"O eroare DB a apărut în timpul importului:\n{type(e).__name__}: {e}"
        logging.error(f"EROARE DB ÎN THREAD-UL DE IMPORT: {e}", exc_info=True)
        if thread_conn_local:
            try: thread_conn_local.rollback()
            except: pass
        q_ref.put(("error", "import_batch", error_message))
    except Exception as e:
        error_message = f"O eroare generală a apărut în timpul importului:\n{type(e).__name__}: {e}"
        logging.error(f"EROARE CRITICĂ ÎN THREAD-UL DE IMPORT: {e}", exc_info=True)
        if thread_conn_local:
            try: thread_conn_local.rollback()
            except: pass
        q_ref.put(("error", "import_batch", error_message))
    finally:
        if thread_conn_local and thread_conn_local.open:
            thread_conn_local.close()

def threaded_export_worker(app_instance, query_str, query_params, file_path_export, q_ref):
    """
    Funcția executată în thread pentru exportul în Excel.
    Versiune finală cu formatare profesională și corecție pentru NameError.
    """
    engine = None
    try:
        db_creds = app_instance.db_handler.db_credentials
        
        db_uri = (
            f"mysql+pymysql://{db_creds['user']}:{db_creds['password']}"
            f"@{db_creds['host']}:{db_creds['port']}/{db_creds['database']}?charset=utf8mb4"
        )
        engine = create_engine(db_uri)

        q_ref.put(("status", "Se preiau datele din baza de date..."))
        df = pd.read_sql_query(query_str, engine, params=query_params)
        df.fillna('', inplace=True)

        if 'data' in df.columns:
            df['data'] = pd.to_datetime(df['data']).dt.date
        
        q_ref.put(("status", "Se scrie și se formatează fișierul Excel..."))

        with pd.ExcelWriter(file_path_export, engine='openpyxl', date_format='YYYY-MM-DD') as writer:
            df.to_excel(writer, index=False, sheet_name='Tranzactii')
            ws = writer.sheets['Tranzactii']

            # Definirea Stilurilor
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            body_font = Font(name='Calibri', size=11)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            credit_fill = PatternFill(start_color='E6F4EA', fill_type='solid')
            debit_fill = PatternFill(start_color='FDE6E6', fill_type='solid')
            
            align_right = Alignment(horizontal='right', vertical='center')
            align_left_vertical_center = Alignment(horizontal='left', vertical='center')

            try:
                tip_col_index = df.columns.get_loc('tip') + 1
            except KeyError:
                tip_col_index = -1

            # Aplicarea Formatelor
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
                is_header = (row_idx == 1)
                
                row_fill = None
                if not is_header and tip_col_index != -1:
                    cell_value = ws.cell(row=row_idx, column=tip_col_index).value
                    if cell_value == 'credit':
                        row_fill = credit_fill
                    elif cell_value == 'debit':
                        row_fill = debit_fill
                
                for cell in row:
                    cell.border = thin_border
                    if is_header:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    else:
                        cell.font = body_font
                        if row_fill:
                            cell.fill = row_fill
                        
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_left_vertical_center

            # Ajustarea Automată a Lățimii Coloanelor
            for i, column_cells in enumerate(ws.columns, start=1):
                # Inițializăm cu lungimea header-ului, care este adesea relevantă
                max_length = len(str(ws.cell(row=1, column=i).value))
                column_letter = get_column_letter(i)
                
                # === AICI ESTE LINIA CORECTATĂ ===
                # Folosim 'column_cells' (plural), așa cum a fost definit în bucla for
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) if max_length < 48 else 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Activarea Funcționalităților Avansate
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

        q_ref.put(("done", "export", file_path_export))
        
    except Exception as e_export:
        logging.error(f"EROARE EXPORT EXCEL: {e_export}", exc_info=True)
        q_ref.put(("error", "export", f"Eroare la exportul în Excel:\n{e_export}"))
    finally:
        if engine:
            engine.dispose()

def threaded_export_to_memory_worker(db_credentials, query_str, query_params):
    """
    Funcție nouă, adaptată. Generează un fișier Excel în memorie.
    Returnează un tuplu: (success: bool, result: BytesIO sau str).
    """
    engine = None
    try:
        db_uri = (
            f"mysql+pymysql://{db_credentials['user']}:{db_credentials['password']}"
            f"@{db_credentials['host']}:{db_credentials['port']}/{db_credentials['database']}?charset=utf8mb4"
        )
        engine = create_engine(db_uri)

        df = pd.read_sql_query(query_str, engine, params=query_params)
        df.fillna('', inplace=True)

        if 'data' in df.columns:
            df['data'] = pd.to_datetime(df['data']).dt.date
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl', date_format='YYYY-MM-DD') as writer:
            df.to_excel(writer, index=False, sheet_name='Tranzactii')
            ws = writer.sheets['Tranzactii']

            # Logica de formatare este identică cu funcția originală
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            body_font = Font(name='Calibri', size=11)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            credit_fill = PatternFill(start_color='E6F4EA', fill_type='solid')
            debit_fill = PatternFill(start_color='FDE6E6', fill_type='solid')
            align_right = Alignment(horizontal='right', vertical='center')
            align_left_vertical_center = Alignment(horizontal='left', vertical='center')

            try:
                tip_col_index = df.columns.get_loc('tip') + 1
            except KeyError:
                tip_col_index = -1

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
                is_header = (row_idx == 1)
                row_fill = None
                if not is_header and tip_col_index != -1:
                    cell_value = ws.cell(row=row_idx, column=tip_col_index).value
                    if cell_value == 'credit': row_fill = credit_fill
                    elif cell_value == 'debit': row_fill = debit_fill
                
                for cell in row:
                    cell.border = thin_border
                    if is_header:
                        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment
                    else:
                        cell.font = body_font
                        if row_fill: cell.fill = row_fill
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_left_vertical_center
            
            for i, column_cells in enumerate(ws.columns, start=1):
                max_length = len(str(ws.cell(row=1, column=i).value))
                column_letter = get_column_letter(i)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                adjusted_width = (max_length + 2) if max_length < 48 else 50
                ws.column_dimensions[column_letter].width = adjusted_width

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

        # Mutăm cursorul la începutul buffer-ului pentru a putea fi citit
        excel_buffer.seek(0)
        return True, excel_buffer

    except Exception as e:
        logging.error(f"EROARE la generarea Excel în memorie: {e}", exc_info=True)
        return False, f"Eroare la generarea fișierului Excel:\n{e}"
    finally:
        if engine: engine.dispose()