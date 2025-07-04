# ui_reports.py
import io
import os
import logging
import numpy as np
import matplotlib.pyplot as plt
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog, Listbox, MULTIPLE, END, Scrollbar
from tkcalendar import DateEntry
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import calendar
import pandas as pd
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl.styles import Font as ExcelFont, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import matplotlib.ticker as mticker
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as ReportlabImage, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

from .email_handler import send_report_email
# Modulul de constante este acum în 'common'
from common.app_constants import APP_NAME, APP_VERSION

# --- Constante pentru stilizarea graficului ---
DIALOG_FONT_FAMILY = 'TkDefaultFont'
DIALOG_FONT_SIZE = 10
CHART_FONT_FAMILY = 'Segoe UI'
CHART_TITLE_FONT_SIZE = 12
CHART_LABEL_FONT_SIZE = 10
CHART_TICK_FONT_SIZE = 9
CHART_TEXT_COLOR = '#333333'
CHART_GRID_COLOR = '#cccccc'
CHART_SPINE_COLOR = '#bbbbbb'
CHART_BG_COLOR = '#f0f0f0'

class CashFlowReportDialog(tk.Toplevel):
    def __init__(self, parent, db_handler, accounts_list, initial_context=None, smtp_config=None):
        super().__init__(parent)
        self.db_handler = db_handler
        self.all_accounts = accounts_list
        self.initial_context = initial_context or {}
        self.visible_tx_codes = self.initial_context.get('visible_tx_codes', [])
        
        # --- NOU: Salvăm permisiunea de acces la tranzacții ---
        self.tranzactie_acces = self.initial_context.get('tranzactie_acces', 'toate')
        
        self.smtp_config = smtp_config or {}
        
        self.title("Analiză Flux de Numerar (Cash Flow)")
        self.minsize(950, 650)
        self.transient(parent)
        
        self.period_var = tk.StringVar()
        self.account_var = tk.StringVar()
        self._report_job = None

        self._create_widgets()
        self._apply_initial_context()
        self.center_window()
        self.grab_set()
        
        self.after(50, self._on_generate_report)

    def center_window(self):
        self.update_idletasks()
        dialog_width = self.winfo_width()
        dialog_height = self.winfo_height()
        parent_x = self.master.winfo_x()
        parent_y = self.master.winfo_y()
        parent_width = self.master.winfo_width()
        parent_height = self.master.winfo_height()
        position_x = parent_x + (parent_width // 2) - (dialog_width // 2)
        position_y = parent_y + (parent_height // 2) - (dialog_height // 2)
        self.geometry(f"+{max(0, position_x)}+{max(0, position_y)}")

    def _generate_pdf_file(self, output_path):
        doc = SimpleDocTemplate(output_path, pagesize=A4, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
        styles = getSampleStyleSheet()
        story = []
        title_text = f"Raport Flux de Numerar: {self.account_var.get()} ({self.current_report_currency})"
        story.append(Paragraph(title_text, styles['h1']))
        story.append(Spacer(1, 0.2*inch))
        img_buffer = io.BytesIO()
        self.fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
        img_buffer.seek(0)
        story.append(ReportlabImage(img_buffer, width=7*inch, height=3.5*inch))
        story.append(Spacer(1, 0.3*inch))
        header_text = "Data" if self.current_report_granularity == 'daily' else "Luna / Anul"
        table_data = [[header_text, 'Total Intrări', 'Total Ieșiri', 'Sold Net']]
        for row in self.current_report_data:
            label = f"{calendar.month_name[row['luna']].capitalize()} {row['an']}" if self.current_report_granularity == 'monthly' else row['data'].strftime('%d-%m-%Y')
            table_data.append([
                Paragraph(label, styles['Normal']),
                f"{float(row['total_intrari']):,.2f}",
                f"{float(row['total_iesiri']):,.2f}",
                f"{float(row['total_intrari']) - float(row['total_iesiri']):,.2f}"
            ])
        table_data.append([
            Paragraph('<b>TOTAL PERIOADĂ</b>', styles['Normal']),
            f"{self.current_report_totals['intrari']:,.2f}",
            f"{self.current_report_totals['iesiri']:,.2f}",
            f"{self.current_report_totals['sold']:,.2f}"
        ])
        table = Table(table_data, colWidths=[2.5*inch, 1.7*inch, 1.7*inch, 1.7*inch])
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'), ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('BOTTOMPADDING', (0,0), (-1,0), 10),
            ('TOPPADDING', (0,0), (-1,0), 10), ('BOTTOMPADDING', (0,-1), (-1,-1), 10),
            ('TOPPADDING', (0,-1), (-1,-1), 10), ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ])
        table.setStyle(style)
        story.append(table)
        doc.build(story)

    def export_to_pdf(self):
        if not hasattr(self, 'current_report_data') or not self.current_report_data:
            messagebox.showwarning("Export Anulat", "Nu există date de exportat.", parent=self)
            return
        default_filename = f"Raport_CashFlow_{self.account_var.get()}_{date.today().strftime('%Y%m%d')}.pdf"
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Fișiere PDF", "*.pdf")], title="Salvează Raportul PDF", initialfile=default_filename)
        if not file_path: return
        try:
            self._generate_pdf_file(file_path)
            messagebox.showinfo("Succes", f"Raportul a fost salvat cu succes în:\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Export PDF", f"A apărut o eroare la salvarea fișierului PDF:\n{e}", parent=self)
            logging.error(f"Eroare la export PDF (CashFlow): {e}", exc_info=True)

    def export_to_excel(self):
        if not hasattr(self, 'current_report_data') or not self.current_report_data:
            messagebox.showwarning("Export Anulat", "Nu există date de exportat.", parent=self)
            return

        default_filename = f"Raport_CashFlow_{self.account_var.get().replace(' ', '_')}_{date.today().strftime('%Y%m%d')}.xlsx"
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Fișiere Excel", "*.xlsx")], title="Salvează Raportul Excel", initialfile=default_filename)
        if not file_path: return

        try:
            header_text = "Data" if self.current_report_granularity == 'daily' else "Perioada"
            df_data = []
            for row in self.current_report_data:
                label = row['data'].strftime('%d.%m.%Y') if self.current_report_granularity == 'daily' else f"{calendar.month_name[row['luna']].capitalize()} {row['an']}"
                intrari = float(row['total_intrari'])
                iesiri = float(row['total_iesiri'])
                df_data.append({
                    header_text: label,
                    'Total Intrări': intrari,
                    'Total Ieșiri': iesiri,
                    'Sold Net': intrari - iesiri
                })
            
            # Adăugăm și linia de total
            df_data.append({
                header_text: 'TOTAL PERIOADĂ',
                'Total Intrări': self.current_report_totals['intrari'],
                'Total Ieșiri': self.current_report_totals['iesiri'],
                'Sold Net': self.current_report_totals['sold']
            })

            df = pd.DataFrame(df_data)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                sheet_name = "Analiza Cash Flow"
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                ws = writer.sheets[sheet_name]

                # Stiluri
                title_font = ExcelFont(bold=True, size=14, name='Calibri')
                header_font = ExcelFont(bold=True, color="FFFFFF", name='Calibri', size=11)
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                total_font = ExcelFont(bold=True, name='Calibri', size=11)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Titlu
                title_cell = ws['A1']
                title_cell.value = f"Raport Flux de Numerar: {self.account_var.get()} ({self.current_report_currency})"
                title_cell.font = title_font
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                title_cell.alignment = Alignment(horizontal='center')

                # Formatare Header
                for cell in ws[2]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                
                # Formatare celule și lățime coloane
                ws.column_dimensions['A'].width = 25
                for col_letter in ['B', 'C', 'D']:
                    ws.column_dimensions[col_letter].width = 18

                for row_cells in ws.iter_rows(min_row=3, max_row=len(df)+1, min_col=1, max_col=4):
                    for cell in row_cells:
                        cell.border = thin_border
                        if cell.column > 1:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                
                # Formatare rând total
                for cell in ws[len(df)+1]:
                    cell.font = total_font

                # Inserare grafic
                img_buffer = io.BytesIO()
                self.fig.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
                img = OpenpyxlImage(img_buffer)
                img.anchor = 'F2'
                ws.add_image(img)

            messagebox.showinfo("Succes", f"Raportul a fost salvat cu succes în:\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Export Excel", f"A apărut o eroare la salvarea fișierului:\n{e}", parent=self)
            logging.error(f"Eroare la export Excel (CashFlow): {e}", exc_info=True)


    def _on_send_email(self):
        if not hasattr(self, 'current_report_data') or not self.current_report_data:
            messagebox.showwarning("Acțiune Anulată", "Vă rugăm mai întâi generați un raport.", parent=self)
            return

        recipient = simpledialog.askstring("Adresă Destinatar", "Introduceți adresa de email a destinatarului:", parent=self)
        if not recipient:
            return

        try:
            account_name = self.account_var.get()
            currency = self.current_report_currency
            start_date_obj, end_date_obj = self.start_date_entry.get_date(), self.end_date_entry.get_date()
            start_date_str, end_date_str = start_date_obj.strftime('%d.%m.%Y'), end_date_obj.strftime('%d.%m.%Y')
            total_intrari, total_iesiri, sold_net = self.current_report_totals['intrari'], self.current_report_totals['iesiri'], self.current_report_totals['sold']

            f_intrari = f"{total_intrari:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            f_iesiri = f"{total_iesiri:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            f_sold = f"{sold_net:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            sign = '+' if sold_net >= 0 else ''
        except Exception as e:
            messagebox.showerror("Eroare Preluare Date", f"Nu s-au putut prelua datele pentru email: {e}", parent=self)
            return

        subject = f"Raport Flux de Numerar: {account_name} ({start_date_str} - {end_date_str})"
        body = f"""Bună ziua,

    Atașat găsiți raportul de flux de numerar generat din aplicația {APP_NAME}.

    --- SUMAR RAPORT ---
    Perioada Raportată: {start_date_str} - {end_date_str}
    Cont Bancar: {account_name}
    Valută: {currency}
    --------------------
    Total Intrări: +{f_intrari} {currency}
    Total Ieșiri: -{f_iesiri} {currency}
    Sold Net: {sign}{f_sold} {currency}
    --------------------

    O zi bună,
    Email generat automat de {APP_NAME} v{APP_VERSION}
    """

        temp_dir = tempfile.gettempdir()
        temp_pdf_path = os.path.join(temp_dir, f"Raport_CashFlow_{account_name.replace(' ', '_')}_{date.today().strftime('%Y%m%d')}.pdf")
        
        try:
            self._generate_pdf_file(temp_pdf_path)
            success, message = send_report_email(self.smtp_config, recipient, subject, body, temp_pdf_path)

            if success:
                messagebox.showinfo("Succes", message, parent=self)
            else:
                messagebox.showerror("Eroare Trimitere", message, parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Generare PDF", f"Nu s-a putut genera fișierul PDF pentru atașare:\n{e}", parent=self)
            logging.error(f"Eroare la generare PDF pentru email (CashFlow): {e}", exc_info=True)
        finally:
            if os.path.exists(temp_pdf_path):
                try:
                    os.remove(temp_pdf_path)
                except OSError as e:
                    logging.warning(f"Nu s-a putut șterge fișierul PDF temporar: {temp_pdf_path}. Eroare: {e}")

    def _apply_initial_context(self):
        logging.debug("DEBUG_CONTEXT: Se aplică contextul inițial în filtrele raportului.")
        initial_account_id = self.initial_context.get('active_account_id')
        initial_account_obj = next((acc for acc in self.all_accounts if acc['id_cont'] == initial_account_id), None)
        if initial_account_obj:
            self.account_var.set(initial_account_obj['nume_cont'])
        elif self.all_accounts:
            self.account_var.set(self.all_accounts[0]['nume_cont'])
        initial_start_date = self.initial_context.get('start_date')
        initial_end_date = self.initial_context.get('end_date')
        if initial_start_date and initial_end_date:
            self.period_var.set("Interval Personalizat")
            self._on_period_select()
            self.start_date_entry.set_date(initial_start_date)
            self.end_date_entry.set_date(initial_end_date)
            logging.debug(f"DEBUG_CONTEXT: Datele raportului setate la {initial_start_date} -> {initial_end_date}")
        else:
            self.period_var.set("Anul Curent")
            self._on_period_select()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)
        top_controls_frame = ttk.LabelFrame(main_frame, text="Filtre și Acțiuni", padding="10")
        top_controls_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        filters_frame = ttk.Frame(top_controls_frame)
        filters_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(filters_frame, text="Cont:").pack(side=tk.LEFT, padx=(0,5))
        account_names = [acc['nume_cont'] for acc in self.all_accounts]
        account_combo = ttk.Combobox(filters_frame, textvariable=self.account_var, values=account_names, state="readonly", width=25, font=(DIALOG_FONT_FAMILY, DIALOG_FONT_SIZE))
        account_combo.pack(side=tk.LEFT, padx=(0,10))
        account_combo.bind("<<ComboboxSelected>>", self._regenerate_report_on_event)
        ttk.Label(filters_frame, text="Perioadă:").pack(side=tk.LEFT)
        period_options = ["Interval Personalizat", "Luna Curentă", "Luna Trecută", "Anul Curent", "Anul Trecut", "Ultimele 12 Luni"]
        period_combo = ttk.Combobox(filters_frame, textvariable=self.period_var, values=period_options, state="readonly", width=18, font=(DIALOG_FONT_FAMILY, DIALOG_FONT_SIZE))
        period_combo.pack(side=tk.LEFT, padx=5)
        period_combo.bind("<<ComboboxSelected>>", self._on_period_select)
        period_combo.bind("<<ComboboxSelected>>", self._regenerate_report_on_event, add='+')
        self.start_date_entry = DateEntry(filters_frame, date_pattern='yyyy-mm-dd', width=12, state='disabled', font=(DIALOG_FONT_FAMILY, DIALOG_FONT_SIZE))
        self.start_date_entry.pack(side=tk.LEFT, padx=(5,2))
        self.start_date_entry.bind("<<DateEntrySelected>>", self._regenerate_report_on_event)
        self.end_date_entry = DateEntry(filters_frame, date_pattern='yyyy-mm-dd', width=12, state='disabled', font=(DIALOG_FONT_FAMILY, DIALOG_FONT_SIZE))
        self.end_date_entry.pack(side=tk.LEFT, padx=(2,5))
        self.end_date_entry.bind("<<DateEntrySelected>>", self._regenerate_report_on_event)
        buttons_frame = ttk.Frame(top_controls_frame)
        buttons_frame.pack(side=tk.RIGHT, fill=tk.X)
        self.export_excel_button = ttk.Button(buttons_frame, text="Export Excel", command=self.export_to_excel, state="disabled")
        self.export_excel_button.pack(side=tk.LEFT, padx=(10, 5))
        self.export_pdf_button = ttk.Button(buttons_frame, text="Export PDF", command=self.export_to_pdf, state="disabled")
        self.export_pdf_button.pack(side=tk.LEFT, padx=5)
        self.send_email_button = ttk.Button(buttons_frame, text="Trimite pe Email", command=self._on_send_email, state="disabled")
        self.send_email_button.pack(side=tk.LEFT, padx=5)
        results_notebook = ttk.Notebook(main_frame, style="Custom.TNotebook")
        results_notebook.grid(row=1, column=0, sticky="nsew", padx=5, pady=10)
        table_tab = ttk.Frame(results_notebook); results_notebook.add(table_tab, text="Sumar Tabelar")
        self.totals_summary_frame = ttk.Frame(table_tab, padding="5"); self.totals_summary_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5,0))
        tree_container_frame = ttk.Frame(table_tab); tree_container_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(tree_container_frame, columns=("luna", "intrari", "iesiri", "sold"), show="headings"); self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar = ttk.Scrollbar(tree_container_frame, orient=tk.VERTICAL, command=self.tree.yview); tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y); self.tree.configure(yscrollcommand=tree_scrollbar.set)
        self.tree.heading("luna", text="Luna / Anul"); self.tree.heading("intrari", text="Total Intrări"); self.tree.heading("iesiri", text="Total Ieșiri"); self.tree.heading("sold", text="Sold Net")
        self.tree.column("luna", width=150, anchor="w"); self.tree.column("intrari", width=150, anchor="e"); self.tree.column("iesiri", width=150, anchor="e"); self.tree.column("sold", width=150, anchor="e")
        self.tree.tag_configure('credit', foreground='#006400'); self.tree.tag_configure('debit', foreground='#8B0000')
        total_font = (DIALOG_FONT_FAMILY, DIALOG_FONT_SIZE, 'bold')
        ttk.Label(self.totals_summary_frame, text="Total Intrări:", font=total_font).pack(side=tk.LEFT, padx=(10, 2))
        self.grand_total_credit_label = ttk.Label(self.totals_summary_frame, text="0.00", font=total_font, foreground="#006400"); self.grand_total_credit_label.pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(self.totals_summary_frame, text="Total Ieșiri:", font=total_font).pack(side=tk.LEFT, padx=(10, 2))
        self.grand_total_debit_label = ttk.Label(self.totals_summary_frame, text="0.00", font=total_font, foreground="#8B0000"); self.grand_total_debit_label.pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(self.totals_summary_frame, text="Sold Net:", font=total_font).pack(side=tk.LEFT, padx=(10, 2))
        self.grand_total_sold_label = ttk.Label(self.totals_summary_frame, text="0.00", font=total_font); self.grand_total_sold_label.pack(side=tk.LEFT, padx=(0, 10))
        chart_tab = ttk.Frame(results_notebook); results_notebook.add(chart_tab, text="Grafic Flux Numerar")
        self.fig = Figure(figsize=(8, 4), dpi=100); self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_tab); self.canvas.draw(); self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    def _regenerate_report_on_event(self, event=None):
        if hasattr(self, '_report_job') and self._report_job:
            self.after_cancel(self._report_job)
        self._report_job = self.after(300, self._on_generate_report)

    def _on_period_select(self, event=None):
        if self.period_var.get() == "Interval Personalizat":
            self.start_date_entry.config(state="normal")
            self.end_date_entry.config(state="normal")
        else:
            self.start_date_entry.config(state="disabled")
            self.end_date_entry.config(state="disabled")
    
    def _on_generate_report(self):
        logging.debug("DEBUG_REPORT: Se generează raportul (logică granularitate)...")
        selected_account_name = self.account_var.get()
        if not selected_account_name:
            if self.all_accounts: self.account_var.set(self.all_accounts[0]['nume_cont'])
            else: messagebox.showwarning("Selecție Invalidă", "Niciun cont disponibil.", parent=self); return
        selected_account = next((acc for acc in self.all_accounts if acc['nume_cont'] == self.account_var.get()), None)
        if not selected_account:
            messagebox.showerror("Eroare", "Detaliile contului selectat nu au putut fi găsite.", parent=self); return
        selected_account_id = selected_account['id_cont']
        account_currency = selected_account.get('valuta', 'RON')
        start_date, end_date = None, None
        
        period = self.period_var.get(); today = date.today()
        if period == "Luna Curentă": start_date, end_date = today.replace(day=1), (today.replace(day=1) + relativedelta(months=1)) - relativedelta(days=1)
        elif period == "Luna Trecută": last_month_end = today.replace(day=1) - relativedelta(days=1); start_date, end_date = last_month_end.replace(day=1), last_month_end
        elif period == "Anul Curent": start_date, end_date = today.replace(day=1, month=1), today.replace(day=31, month=12)
        elif period == "Anul Trecut": last_year = today.year - 1; start_date, end_date = date(last_year, 1, 1), date(last_year, 12, 31)
        elif period == "Ultimele 12 Luni": start_date, end_date = today - relativedelta(months=12), today
        elif period == "Interval Personalizat":
            try:
                start_date, end_date = self.start_date_entry.get_date(), self.end_date_entry.get_date()
                if start_date > end_date: messagebox.showerror("Eroare Interval", "Data de început nu poate fi după data de sfârșit.", parent=self); return
            except Exception as e: messagebox.showerror("Eroare Dată", f"Data introdusă este invalidă: {e}", parent=self); return
        
        if not start_date or not end_date:
             # Fallback la contextul inițial dacă există, altfel eroare
            initial_start_date = self.initial_context.get('start_date')
            initial_end_date = self.initial_context.get('end_date')
            if initial_start_date and initial_end_date:
                start_date, end_date = initial_start_date, initial_end_date
            else:
                messagebox.showerror("Eroare", "Intervalul de date nu a putut fi determinat.", parent=self); return

        duration_days = (end_date - start_date).days
        granularity = 'daily' if duration_days <= 62 else 'monthly'
        
        self.start_date_entry.set_date(start_date)
        self.end_date_entry.set_date(end_date)
        
        # --- NOU: Construim clauza SQL pentru accesul la tranzacții ---
        access_sql, access_params = "", []
        if self.tranzactie_acces == 'credit':
            access_sql = " AND tip = %s "
            access_params.append('credit')
        elif self.tranzactie_acces == 'debit':
            access_sql = " AND tip = %s "
            access_params.append('debit')

        base_query = f"FROM tranzactii WHERE id_cont_fk = %s AND data BETWEEN %s AND %s {access_sql}"
        
        if granularity == 'monthly':
            sql = "SELECT YEAR(data) as an, MONTH(data) as luna, " \
                  "SUM(CASE WHEN tip = 'credit' THEN suma ELSE 0 END) as total_intrari, " \
                  f"SUM(CASE WHEN tip = 'debit' THEN suma ELSE 0 END) as total_iesiri {base_query}"
        else: # 'daily'
            sql = "SELECT data, " \
                  "SUM(CASE WHEN tip = 'credit' THEN suma ELSE 0 END) as total_intrari, " \
                  f"SUM(CASE WHEN tip = 'debit' THEN suma ELSE 0 END) as total_iesiri {base_query}"

        params = [selected_account_id, start_date, end_date] + access_params
        
        if self.visible_tx_codes:
            placeholders = ', '.join(['%s'] * len(self.visible_tx_codes))
            sql += f" AND cod_tranzactie_fk IN ({placeholders})"
            params.extend(self.visible_tx_codes)
        
        sql += " GROUP BY 1, 2 ORDER BY 1, 2;" if granularity == 'monthly' else " GROUP BY 1 ORDER BY 1;"
        
        results = self.db_handler.fetch_all_dict(sql, tuple(params))
        grand_totals = { 'intrari': 0.0, 'iesiri': 0.0, 'sold': 0.0 }
        if results:
            grand_totals['intrari'] = sum(float(r['total_intrari']) for r in results)
            grand_totals['iesiri'] = sum(float(r['total_iesiri']) for r in results)
            grand_totals['sold'] = grand_totals['intrari'] - grand_totals['iesiri']
        
        self.current_report_data = results
        self.current_report_totals = grand_totals
        self.current_report_granularity = granularity
        self.current_report_currency = account_currency
        
        self.export_excel_button.config(state="normal" if results else "disabled")
        self.export_pdf_button.config(state="normal" if results else "disabled")
        self.send_email_button.config(state="normal" if results else "disabled")
        
        self._populate_table(results, account_currency, granularity, grand_totals)
        self._update_chart(results, account_currency, granularity)

    def _populate_table(self, data, currency, granularity, grand_totals):
        for i in self.tree.get_children(): self.tree.delete(i)
        header_text = "Data" if granularity == 'daily' else "Luna / Anul"
        self.tree.heading("luna", text=header_text)
        if data:
            for row in data:
                if granularity == 'monthly': label = f"{calendar.month_name[row['luna']].capitalize()} {row['an']}"
                else: label = row['data'].strftime('%d-%m-%Y')
                intrari = float(row['total_intrari']); iesiri = float(row['total_iesiri']); sold = intrari - iesiri
                intrari_str = f"{intrari:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                iesiri_str = f"{iesiri:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                sold_str = f"{sold:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                tags = ('credit',) if sold >= 0 else ('debit',)
                self.tree.insert("", "end", values=(label, intrari_str, iesiri_str, sold_str), tags=tags)
        total_intrari_str = f"{grand_totals['intrari']:,.2f} {currency}".replace(",", "X").replace(".", ",").replace("X", ".")
        total_iesiri_str = f"{grand_totals['iesiri']:,.2f} {currency}".replace(",", "X").replace(".", ",").replace("X", ".")
        total_sold_str = f"{grand_totals['sold']:,.2f} {currency}".replace(",", "X").replace(".", ",").replace("X", ".")
        self.grand_total_credit_label.config(text=total_intrari_str)
        self.grand_total_debit_label.config(text=total_iesiri_str)
        self.grand_total_sold_label.config(text=total_sold_str)
        sold_color = "black"
        if grand_totals['sold'] > 0: sold_color = "#006400"
        elif grand_totals['sold'] < 0: sold_color = "#8B0000"
        self.grand_total_sold_label.config(foreground=sold_color)

    def _update_chart(self, data, currency, granularity):
        self.ax.clear()
        self.ax.set_facecolor(CHART_BG_COLOR)
        self.fig.set_facecolor(CHART_BG_COLOR)
        if not data:
            self.ax.text(0.5, 0.5, 'Nicio dată de afișat.', horizontalalignment='center', verticalalignment='center', fontfamily=CHART_FONT_FAMILY)
            self.canvas.draw()
            return
        if granularity == 'monthly':
            labels = [f"{calendar.month_abbr[r['luna']]} '{str(r['an'])[-2:]}" for r in data]
        else:
            labels = [r['data'].strftime('%d-%m') for r in data]
        intrari = [float(r['total_intrari']) for r in data]
        iesiri = [float(r['total_iesiri']) for r in data]
        x = np.arange(len(labels))
        width = 0.35
        self.ax.bar(x - width/2, intrari, width, label='Intrări', color='#2ca02c')
        self.ax.bar(x + width/2, iesiri, width, label='Ieșiri', color='#d62728')
        title_font = {'family': CHART_FONT_FAMILY, 'size': CHART_TITLE_FONT_SIZE, 'weight': 'bold', 'color': CHART_TEXT_COLOR}
        label_font = {'family': CHART_FONT_FAMILY, 'size': CHART_LABEL_FONT_SIZE, 'color': CHART_TEXT_COLOR}
        self.ax.set_title(f'Flux de Numerar ({granularity.capitalize()}) - {currency}', fontdict=title_font, pad=20)
        self.ax.set_ylabel(f'Sumă ({currency})', fontdict=label_font, labelpad=10)
        self.ax.grid(axis='y', linestyle='--', color=CHART_GRID_COLOR, alpha=0.7)
        self.ax.set_axisbelow(True)
        self.ax.spines['top'].set_visible(False); self.ax.spines['right'].set_visible(False)
        self.ax.spines['left'].set_color(CHART_SPINE_COLOR); self.ax.spines['bottom'].set_color(CHART_SPINE_COLOR)
        self.ax.tick_params(axis='x', colors=CHART_TEXT_COLOR, labelsize=CHART_TICK_FONT_SIZE)
        self.ax.tick_params(axis='y', colors=CHART_TEXT_COLOR, labelsize=CHART_TICK_FONT_SIZE)
        self.ax.set_xticks(x)
        self.ax.set_xticklabels(labels, rotation=45, ha="right", fontfamily=CHART_FONT_FAMILY)
        formatter = mticker.FuncFormatter(lambda val, pos: f'{int(val):,}')
        self.ax.yaxis.set_major_formatter(formatter)
        legend = self.ax.legend(loc='upper right')
        for text in legend.get_texts():
            text.set_fontfamily(CHART_FONT_FAMILY); text.set_color(CHART_TEXT_COLOR); text.set_fontsize(CHART_TICK_FONT_SIZE)
        self.fig.tight_layout()
        self.canvas.draw()

class BalanceEvolutionReportDialog(tk.Toplevel):
    def __init__(self, parent, db_handler, smtp_config, report_config):
        super().__init__(parent)
        self.db_handler = db_handler
        self.smtp_config = smtp_config or {}
        self.report_config = report_config
        
        # --- NOU: Salvăm permisiunea de acces la tranzacții ---
        self.tranzactie_acces = self.report_config.get('tranzactie_acces', 'toate')
        
        self.report_data = []
        self.title(f"Evoluție Sold - {self.report_config['account_name']}")
        self.geometry("1000x700")
        self.transient(parent)
        self._create_widgets()
        self.center_window()
        self.grab_set()
        self.after(50, self._generate_report)

    def center_window(self):
        self.update_idletasks()
        dialog_width = self.winfo_width()
        dialog_height = self.winfo_height()
        parent_x = self.master.winfo_x()
        parent_y = self.master.winfo_y()
        parent_width = self.master.winfo_width()
        parent_height = self.master.winfo_height()
        position_x = parent_x + (parent_width // 2) - (dialog_width // 2)
        position_y = parent_y + (parent_height // 2) - (dialog_height // 2)
        self.geometry(f"+{max(0, position_x)}+{max(0, position_y)}")

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        chart_frame = ttk.Frame(main_frame)
        chart_frame.grid(row=0, column=0, sticky="nsew")
        self.fig = Figure(figsize=(10, 6), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        buttons_frame = ttk.Frame(main_frame, padding=(0, 10, 0, 0))
        buttons_frame.grid(row=1, column=0, sticky="ew")
        self.export_excel_button = ttk.Button(buttons_frame, text="Export Excel", command=self.export_to_excel, state="disabled")
        self.export_excel_button.pack(side=tk.LEFT, padx=5)
        self.export_pdf_button = ttk.Button(buttons_frame, text="Export PDF", command=self.export_to_pdf, state="disabled")
        self.export_pdf_button.pack(side=tk.LEFT, padx=5)
        self.send_email_button = ttk.Button(buttons_frame, text="Trimite pe Email", command=self._on_send_email, state="disabled")
        self.send_email_button.pack(side=tk.LEFT, padx=5)
    
    def _generate_report(self):
        start_date = self.report_config['start_date']
        end_date = self.report_config['end_date']
        account_id = self.report_config['account_id']
        granularity = self.report_config['granularity']
        visible_tx_codes = self.report_config.get('visible_tx_codes', [])

        # --- NOU: Construim clauzele SQL pentru acces și vizibilitate coduri ---
        access_sql, access_params = "", []
        if self.tranzactie_acces == 'credit':
            access_sql = " AND tip = %s "
            access_params.append('credit')
        elif self.tranzactie_acces == 'debit':
            access_sql = " AND tip = %s "
            access_params.append('debit')

        visibility_sql = ""
        visibility_params = []
        if visible_tx_codes:
            placeholders = ', '.join(['%s'] * len(visible_tx_codes))
            visibility_sql = f" AND cod_tranzactie_fk IN ({placeholders})"
            visibility_params.extend(visible_tx_codes)

        # Combinăm filtrele
        filter_sql = access_sql + visibility_sql
        
        # Aplicăm filtrele la interogarea pentru soldul inițial
        initial_balance_query = f"""
            SELECT SUM(CASE WHEN tip = 'credit' THEN suma ELSE -suma END)
            FROM tranzactii
            WHERE id_cont_fk = %s AND data < %s
            {filter_sql}
        """
        initial_balance_params = [account_id, start_date] + access_params + visibility_params
        initial_balance_result = self.db_handler.fetch_scalar(initial_balance_query, tuple(initial_balance_params))
        initial_balance = float(initial_balance_result) if initial_balance_result is not None else 0.0
        logging.debug(f"DEBUG_REPORT (filtrat): Sold inițial calculat (înainte de {start_date}): {initial_balance}")

        # Aplicăm filtrele la interogarea pentru tranzacțiile din perioadă
        transactions_query = f"""
            SELECT data, suma, tip
            FROM tranzactii
            WHERE id_cont_fk = %s AND data BETWEEN %s AND %s
            {filter_sql}
            ORDER BY data ASC, id ASC
        """
        transactions_params = [account_id, start_date, end_date] + access_params + visibility_params
        transactions = self.db_handler.fetch_all_dict(transactions_query, tuple(transactions_params))
        
        # ... (restul metodei, care procesează datele, rămâne neschimbat) ...
        if not transactions and initial_balance == 0:
            self.report_data = []
        else:
            all_days = pd.date_range(start=start_date, end=end_date, freq='D')
            daily_balances = pd.Series(index=all_days, dtype=float).fillna(0)
            if transactions:
                df = pd.DataFrame(transactions)
                df['data'] = pd.to_datetime(df['data'])
                df['net_change'] = df.apply(lambda row: float(row['suma']) if row['tip'] == 'credit' else -float(row['suma']), axis=1)
                daily_changes = df.groupby('data')['net_change'].sum()
                daily_balances = daily_balances.add(daily_changes, fill_value=0)
            
            final_daily_balances = daily_balances.cumsum() + initial_balance
            
            if granularity == 'Lună':
                sampled_balances = final_daily_balances.resample('M').last()
            elif granularity == 'Anuală':
                sampled_balances = final_daily_balances.resample('A').last()
            else:
                sampled_balances = final_daily_balances
            
            self.report_data = [
                {'data': index.date(), 'sold_dupa_tranzactie': value}
                for index, value in sampled_balances.items()
            ]

        if self.report_data:
            self.export_excel_button.config(state="normal")
            self.export_pdf_button.config(state="normal")
            self.send_email_button.config(state="normal")
        else:
            self.export_excel_button.config(state="disabled")
            self.export_pdf_button.config(state="disabled")
            self.send_email_button.config(state="disabled")
            
        self._update_chart()

    def _update_chart(self):
        self.ax.clear()
        if not self.report_data:
            self.ax.text(0.5, 0.5, 'Nicio dată de afișat pentru perioada selectată.', horizontalalignment='center', verticalalignment='center')
        else:
            dates = [row['data'] for row in self.report_data]
            balances = [float(row['sold_dupa_tranzactie']) for row in self.report_data]
            self.ax.plot(dates, balances, marker='o', linestyle='-', color='#007acc', markersize=4)
            title = f"Evoluție Sold: {self.report_config['account_name']} ({self.report_config['start_date'].strftime('%d.%m.%Y')} - {self.report_config['end_date'].strftime('%d.%m.%Y')})"
            self.ax.set_title(title, fontsize=14, weight='bold', pad=20)
            self.ax.set_ylabel(f"Sold Cont ({self.report_config['currency']})", fontsize=12)
            self.ax.grid(True, which='both', linestyle='--', linewidth=0.5)
            formatter = mticker.FuncFormatter(lambda x, pos: f'{int(x):,}')
            self.ax.yaxis.set_major_formatter(formatter)
            self.fig.autofmt_xdate(rotation=30, ha='right')
        self.fig.tight_layout()
        self.canvas.draw()

    def export_to_excel(self):
        if not self.report_data:
            messagebox.showwarning("Export Anulat", "Nu există date de exportat.", parent=self)
            return
            
        default_filename = f"Raport_Evolutie_Sold_{self.report_config['account_name'].replace(' ','_')}_{date.today().strftime('%Y%m%d')}.xlsx"
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Fișiere Excel", "*.xlsx")], title="Salvează Raportul Excel", initialfile=default_filename)
        if not file_path: return

        try:
            df_data = [{'Perioada': row['data'].strftime('%d.%m.%Y'), 'Sold Final': float(row['sold_dupa_tranzactie'])} for row in self.report_data]
            df = pd.DataFrame(df_data)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                sheet_name = f"Evolutie Sold {self.report_config['account_name']}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                ws = writer.sheets[sheet_name]
                
                title_font = ExcelFont(bold=True, size=14, name='Calibri'); header_font = ExcelFont(bold=True, color="FFFFFF", name='Calibri', size=11)
                header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                ws.merge_cells('A1:B1'); title_cell = ws['A1']
                title_cell.value = f"Raport Evoluție Sold: {self.report_config['account_name']}"
                title_cell.font = title_font; title_cell.alignment = Alignment(horizontal='center')

                for cell in ws[2]: cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
                ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 20
                for row_cells in ws.iter_rows(min_row=3, max_row=len(df)+2, min_col=1, max_col=2):
                    for cell in row_cells:
                        cell.border = thin_border
                        if cell.column == 2: cell.number_format = '#,##0.00'
                
                img_buffer = io.BytesIO(); self.fig.savefig(img_buffer, format='png', dpi=200); img = OpenpyxlImage(img_buffer)
                img.anchor = 'D2'; ws.add_image(img)
            
            messagebox.showinfo("Succes", f"Raportul a fost salvat cu succes în:\n{file_path}", parent=self)
        except KeyError as e:
            messagebox.showerror("Eroare de Cheie", f"A apărut o eroare la pregătirea datelor pentru export. Cheie lipsă: {e}", parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Export Excel", f"A apărut o eroare la salvarea fișierului:\n{e}", parent=self)
            logging.error(f"Eroare la export Excel (Evolutie Sold): {e}", exc_info=True)


    def _generate_pdf_file(self, output_path):
        doc = SimpleDocTemplate(output_path, pagesize=A4, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
        styles = getSampleStyleSheet()
        story = []
        title_text = f"Raport Evoluție Sold: {self.report_config['account_name']}"
        story.append(Paragraph(title_text, styles['h1']))
        period_text = f"Perioada: {self.report_config['start_date'].strftime('%d.%m.%Y')} - {self.report_config['end_date'].strftime('%d.%m.%Y')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        img_buffer = io.BytesIO(); self.fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight'); img_buffer.seek(0)
        story.append(ReportlabImage(img_buffer, width=7.5*inch, height=5*inch))
        
        story.append(Spacer(1, 0.3*inch))
        table_data = [['Perioada', 'Sold la Sfârșitul Perioadei']]
        for row in self.report_data:
            label = row['data'].strftime('%d.%m.%Y')
            balance_str = f"{float(row['sold_dupa_tranzactie']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            table_data.append([label, balance_str])
        
        table = Table(table_data, colWidths=[3.75*inch, 3.75*inch])
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#007ACC')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ])
        table.setStyle(style)
        story.append(table)
        
        doc.build(story)

    def export_to_pdf(self):
        if not self.report_data: return
        default_filename = f"Raport_Evolutie_Sold_{self.report_config['account_name'].replace(' ','_')}_{date.today().strftime('%Y%m%d')}.pdf"
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Fișiere PDF", "*.pdf")], title="Salvează Raportul PDF", initialfile=default_filename)
        if not file_path: return
        try:
            self._generate_pdf_file(file_path)
            messagebox.showinfo("Succes", f"Raportul a fost salvat cu succes în:\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Export PDF", f"A apărut o eroare la salvarea fișierului PDF:\n{e}", parent=self)
            logging.error(f"Eroare la export PDF (Evolutie Sold): {e}", exc_info=True)


    def _on_send_email(self):
        if not self.report_data: return
        recipient = simpledialog.askstring("Adresă Destinatar", "Introduceți adresa de email a destinatarului:", parent=self)
        if not recipient: return

        temp_dir = tempfile.gettempdir(); temp_pdf_path = os.path.join(temp_dir, f"Evolutie_Sold_{self.report_config['account_name'].replace(' ','_')}.pdf")
        
        try:
            self._generate_pdf_file(temp_pdf_path)
            
            start_date_str = self.report_config['start_date'].strftime('%d.%m.%Y')
            end_date_str = self.report_config['end_date'].strftime('%d.%m.%Y')
            
            start_balance = f"{float(self.report_data[0]['sold_dupa_tranzactie']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            end_balance = f"{float(self.report_data[-1]['sold_dupa_tranzactie']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            currency = self.report_config['currency']

            subject = f"Evoluție Sold Cont: {self.report_config['account_name']} ({start_date_str} - {end_date_str})"
            body = f"""Bună ziua,

Atașat găsiți raportul privind evoluția soldului pentru contul {self.report_config['account_name']}.

--- SUMAR PERIOADĂ ---
Perioada analizată: {start_date_str} - {end_date_str}
Sold de început (aproximat): {start_balance} {currency}
Sold de sfârșit: {end_balance} {currency}
--------------------

O zi bună,
Email generat automat de {APP_NAME} v{APP_VERSION}
"""
            success, message = send_report_email(self.smtp_config, recipient, subject, body, temp_pdf_path)
            if success: messagebox.showinfo("Succes", message, parent=self)
            else: messagebox.showerror("Eroare Trimitere", message, parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Critică", f"A apărut o eroare neașteptată:\n{e}", parent=self)
            logging.error(f"Eroare la trimitere email (Evolutie Sold): {e}", exc_info=True)
        finally:
            if os.path.exists(temp_pdf_path):
                try:
                    os.remove(temp_pdf_path)
                except OSError as e:
                    logging.warning(f"Nu s-a putut șterge fișierul PDF temporar: {temp_pdf_path}. Eroare: {e}")

class TransactionAnalysisReportDialog(tk.Toplevel):
    def __init__(self, parent, db_handler, initial_context=None, smtp_config=None):
        super().__init__(parent)
        self.db_handler = db_handler
        self.initial_context = initial_context or {}
        
        self.tranzactie_acces = self.initial_context.get('tranzactie_acces', 'toate')
        self.smtp_config = smtp_config or self.initial_context.get('smtp_config', {})
            
        self.title("Analiză Detaliată a Tranzacțiilor")
        self.geometry("1100x750")
        self.transient(parent)

        # Variabile de control pentru filtre
        self.granularity_var = tk.StringVar(value="Zilnic")
        self.type_var = tk.StringVar(value="Ambele")
        
        # Atribute interne
        self.report_data = []
        self.transaction_type_vars = {}
        self.legend_desc_map = {}
        self.results_tree = None
        self.canvas = None

        # Inițializăm sistemul de tooltip
        self.tooltip = tk.Toplevel(self)
        self.tooltip.withdraw()
        self.tooltip.overrideredirect(True)
        self.tooltip_label = ttk.Label(self.tooltip, text="", justify='left',
                                    background='#FFFFE0', relief='solid', borderwidth=1,
                                    font=(DIALOG_FONT_FAMILY, 9))
        self.tooltip_label.pack(ipadx=5, ipady=3)
        
        # Construim widget-urile UI
        self._create_widgets()

        # Legăm evenimentul de hover DUPĂ ce self.canvas a fost creat
        if self.canvas:
            self.canvas.mpl_connect('motion_notify_event', self._on_mouse_hover)
        
        # --- BLOCUL CRUCIAL CARE RESTAUREAZĂ FUNCȚIONALITATEA ---
        self._populate_filters_with_initial_context()
        self._setup_bindings() # Asigurăm legarea evenimentelor
        self._schedule_report_update() # Pornim prima actualizare
        # --- SFÂRȘIT BLOC ---

        self.center_window()
        self.grab_set()

    def center_window(self):
        self.update_idletasks()
        dialog_width = self.winfo_width()
        dialog_height = self.winfo_height()
        parent_x = self.master.winfo_x()
        parent_y = self.master.winfo_y()
        parent_width = self.master.winfo_width()
        parent_height = self.master.winfo_height()
        position_x = parent_x + (parent_width // 2) - (dialog_width // 2)
        position_y = parent_y + (parent_height // 2) - (dialog_height // 2)
        self.geometry(f"+{max(0, position_x)}+{max(0, position_y)}")

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.rowconfigure(2, weight=1) 
        main_frame.columnconfigure(0, weight=1)

        filter_panel = ttk.LabelFrame(main_frame, text="Filtre și Opțiuni", padding="10")
        filter_panel.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

        filter_row1 = ttk.Frame(filter_panel)
        filter_row1.pack(fill="x", expand=True, pady=(0, 10))

        ttk.Label(filter_row1, text="De la:").pack(side="left", padx=(0, 5))
        self.start_date_entry = DateEntry(filter_row1, date_pattern='yyyy-mm-dd', width=12)
        self.start_date_entry.pack(side="left", padx=(0, 15))

        ttk.Label(filter_row1, text="Până la:").pack(side="left", padx=(0, 5))
        self.end_date_entry = DateEntry(filter_row1, date_pattern='yyyy-mm-dd', width=12)
        self.end_date_entry.pack(side="left", padx=(0, 15))

        ttk.Label(filter_row1, text="Tip Tranzacție:").pack(side="left", padx=(0, 5))
        self.type_combo = ttk.Combobox(filter_row1, textvariable=self.type_var, values=["Ambele", "Doar Credit", "Doar Debit"], state="readonly", width=12)
        self.type_combo.pack(side="left", padx=(0, 15))

        ttk.Label(filter_row1, text="Granularitate Listă:").pack(side="left", padx=(0, 5))
        self.granularity_combo = ttk.Combobox(filter_row1, textvariable=self.granularity_var, values=["Zilnic", "Lunar", "Anual"], state="readonly", width=12)
        self.granularity_combo.pack(side="left")

        checkbutton_container_frame = ttk.LabelFrame(filter_panel, text="Coduri Tranzacție (Filtru Listă și Grafic)")
        checkbutton_container_frame.pack(fill="x", pady=5)
        
        canvas = tk.Canvas(checkbutton_container_frame, borderwidth=0, background="#ffffff")
        scrollbar = ttk.Scrollbar(checkbutton_container_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self.checkbuttons_frame = ttk.Frame(canvas, padding=5)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        canvas_frame = canvas.create_window((0, 0), window=self.checkbuttons_frame, anchor="nw")
        
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.config(height=self.checkbuttons_frame.winfo_reqheight())
        
        def on_canvas_configure(event):
            canvas.itemconfig(canvas_frame, width=event.width)

        self.checkbuttons_frame.bind("<Configure>", on_frame_configure)
        canvas.bind('<Configure>', on_canvas_configure)

        buttons_frame = ttk.Frame(main_frame, padding=(0, 10, 0, 0))
        buttons_frame.grid(row=1, column=0, sticky="ew", padx=5)

        self.export_excel_button = ttk.Button(buttons_frame, text="Export Excel", command=self.export_to_excel, state="disabled")
        self.export_excel_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.export_pdf_button = ttk.Button(buttons_frame, text="Export PDF", command=self.export_to_pdf, state="disabled")
        self.export_pdf_button.pack(side=tk.LEFT, padx=5)

        self.send_email_button = ttk.Button(buttons_frame, text="Trimite pe Email", command=self._on_send_email, state="disabled")
        self.send_email_button.pack(side=tk.LEFT, padx=5)
        
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=2, column=0, sticky="nsew", padx=5, pady=(5, 0))

        list_tab = ttk.Frame(self.notebook)
        self.notebook.add(list_tab, text=" Listă Detaliată ")
        
        tree_frame = ttk.Frame(list_tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        tree_cols = ("perioada", "cod_tranzactie", "tip", "total_suma", "nr_tranzactii")
        self.results_tree = ttk.Treeview(tree_frame, columns=tree_cols, show="headings")
        self.results_tree.pack(side="left", fill="both", expand=True)
        
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.results_tree.yview)
        tree_scrollbar.pack(side="right", fill="y")
        self.results_tree.configure(yscrollcommand=tree_scrollbar.set)

        self.results_tree.heading("perioada", text="Perioada")
        self.results_tree.heading("cod_tranzactie", text="Cod Tranzacție")
        self.results_tree.heading("tip", text="Tip")
        self.results_tree.heading("total_suma", text="Sumă Totală")
        self.results_tree.heading("nr_tranzactii", text="Nr. Tranzacții")
        
        self.results_tree.column("perioada", width=120, anchor="w")
        self.results_tree.column("cod_tranzactie", width=120, anchor="center")
        self.results_tree.column("tip", width=100, anchor="center")
        self.results_tree.column("total_suma", width=150, anchor="e")
        self.results_tree.column("nr_tranzactii", width=100, anchor="center")

        graph_tab = ttk.Frame(self.notebook)
        self.notebook.add(graph_tab, text=" Grafic Sumar ")
        
        self.fig = Figure(figsize=(10, 6), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=graph_tab)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    def _show_tooltip(self, text, x, y):
        """Afișează tooltip-ul la coordonatele specificate."""
        self.tooltip_label.config(text=text)
        self.tooltip.geometry(f"+{x}+{y}")
        self.tooltip.deiconify()

    def _hide_tooltip(self, event=None):
        """Ascunde tooltip-ul."""
        self.tooltip.withdraw()

    def _schedule_report_update(self, event=None):
        if hasattr(self, '_update_job'):
            self.after_cancel(self._update_job)
        self._update_job = self.after(400, self._refresh_report_data)

    def _setup_bindings(self):
        """Leagă evenimentele widget-urilor de funcția de actualizare."""
        self.start_date_entry.bind("<<DateEntrySelected>>", self._schedule_report_update)
        self.end_date_entry.bind("<<DateEntrySelected>>", self._schedule_report_update)
        self.granularity_combo.bind("<<ComboboxSelected>>", self._schedule_report_update)
        self.type_combo.bind("<<ComboboxSelected>>", self._schedule_report_update)
        # Checkbutton-urile își apelează funcția direct prin parametrul 'command',
        # deci nu au nevoie de .bind() aici.

    def _on_mouse_hover(self, event):
        """Metoda finală și robustă pentru gestionarea tooltip-urilor pe grafic."""
        found_active_element = False
        
        # 1. Verificăm întâi legenda
        legend = self.ax.get_legend()
        if legend:
            # Iterăm prin elementele text ale legendei
            for legtext in legend.get_texts():
                # Folosim metoda robustă .contains() a artistului
                contains, _ = legtext.contains(event)
                if contains:
                    short_code = legtext.get_text()
                    full_description = self.legend_desc_map.get(short_code, "Descriere indisponibilă")
                    self._show_tooltip(full_description, event.x_root + 15, event.y_root + 10)
                    found_active_element = True
                    break # Am găsit, ieșim din bucla legendei
        
        # 2. Dacă nu am găsit nimic în legendă, verificăm barele (doar dacă suntem în zona axelor)
        if not found_active_element and event.inaxes == self.ax:
            for bar in self.ax.patches:
                contains, _ = bar.contains(event)
                if contains:
                    value = bar.get_height()
                    currency = self.initial_context.get('currency', 'RON')
                    tooltip_text = f"{value:,.2f} {currency}".replace(",", "X").replace(".", ",").replace("X", ".")
                    self._show_tooltip(tooltip_text, event.x_root + 15, event.y_root + 10)
                    found_active_element = True
                    break # Am găsit, ieșim din bucla barelor

        # 3. Dacă nu am găsit nimic, ascundem tooltip-ul
        if not found_active_element:
            self._hide_tooltip()

    def _populate_filters_with_initial_context(self):
        """Setează valorile inițiale și creează dinamic Checkbutton-urile."""
        if not self.initial_context:
            return

        start_date = self.initial_context.get('start_date')
        end_date = self.initial_context.get('end_date')
        if start_date: self.start_date_entry.set_date(start_date)
        if end_date: self.end_date_entry.set_date(end_date)

        all_transaction_types = self.db_handler.fetch_all_dict(
            "SELECT cod, descriere_tip FROM tipuri_tranzactii ORDER BY cod ASC"
        )
        visible_codes = self.initial_context.get('visible_tx_codes', [])
        
        active_account_id = self.initial_context.get('active_account_id')
        accounts_list = self.initial_context.get('accounts_list', [])
        active_account = next((acc for acc in accounts_list if acc['id_cont'] == active_account_id), None)
        if active_account:
            self.initial_context['account_name'] = active_account.get('nume_cont', 'N/A')

        max_cols = 5
        row, col = 0, 0

        if all_transaction_types:
            for type_info in all_transaction_types:
                code = type_info['cod']
                description = type_info['descriere_tip']
                is_active = code in visible_codes
                var = tk.BooleanVar(value=is_active)
                
                cb = ttk.Checkbutton(self.checkbuttons_frame, text=code, variable=var, command=self._schedule_report_update)
                cb.grid(row=row, column=col, sticky="w", padx=5, pady=2)
                
                def on_enter(e, text=description):
                    self._show_tooltip(text, e.x_root + 10, e.y_root + 20)
                
                cb.bind("<Enter>", on_enter)
                cb.bind("<Leave>", self._hide_tooltip)

                self.transaction_type_vars[code] = var
                col += 1
                if col >= max_cols:
                    col = 0
                    row += 1

    def _refresh_report_data(self):
        """Construiește query-ul, preia datele din DB și pregătește totul pentru UI."""
        try:
            start_date = self.start_date_entry.get_date()
            end_date = self.end_date_entry.get_date()
        except (ValueError, TypeError):
            messagebox.showerror("Dată Invalidă", "Vă rugăm introduceți un interval de date valid.", parent=self)
            return

        granularity = self.granularity_var.get()
        transaction_type_filter = self.type_var.get()

        selected_codes = [code for code, var in self.transaction_type_vars.items() if var.get()]
        
        if not selected_codes:
            self.report_data = []
            self._populate_treeview()
            self._update_graph() 
            self.export_excel_button.config(state="disabled")
            self.export_pdf_button.config(state="disabled")
            self.send_email_button.config(state="disabled")
            return

        if granularity == "Zilnic": select_period_col, group_by_period = "data", "data"
        elif granularity == "Lunar": select_period_col, group_by_period = "DATE_FORMAT(data, '%Y-%m')", "perioada"
        else: select_period_col, group_by_period = "YEAR(data)", "perioada"

        sql = f"SELECT {select_period_col} as perioada, cod_tranzactie_fk, tip, SUM(suma) as total_suma, COUNT(*) as nr_tranzactii FROM tranzactii WHERE id_cont_fk = %s AND data BETWEEN %s AND %s"
        params = [self.initial_context['active_account_id'], start_date, end_date]
        
        if self.tranzactie_acces == 'credit':
            sql += " AND tip = %s"
            params.append('credit')
        elif self.tranzactie_acces == 'debit':
            sql += " AND tip = %s"
            params.append('debit')

        if transaction_type_filter != "Ambele":
            sql += " AND tip = %s"
            params.append('credit' if transaction_type_filter == "Doar Credit" else 'debit')

        placeholders = ', '.join(['%s'] * len(selected_codes))
        sql += f" AND cod_tranzactie_fk IN ({placeholders})"
        params.extend(selected_codes)
        sql += f" GROUP BY {group_by_period}, cod_tranzactie_fk, tip ORDER BY perioada DESC, cod_tranzactie_fk"

        self.report_data = self.db_handler.fetch_all_dict(sql, tuple(params))
        
        self._populate_treeview()
        self._update_graph()

        final_state = "normal" if self.report_data else "disabled"
        self.export_excel_button.config(state=final_state)
        self.export_pdf_button.config(state=final_state)
        self.send_email_button.config(state=final_state)

    def _populate_treeview(self):
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)

        self.results_tree.tag_configure('credit', foreground='#006400')
        self.results_tree.tag_configure('debit', foreground='#8B0000')

        if self.report_data:
            currency = self.initial_context.get('currency', 'RON')
            for row in self.report_data:
                period = row['perioada']
                if isinstance(period, date):
                    period = period.strftime('%d-%m-%Y')

                total_suma_str = f"{float(row['total_suma']):,.2f} {currency}".replace(",", "X").replace(".", ",").replace("X", ".")
                tag = 'credit' if row['tip'] == 'credit' else 'debit'
                values = (period, row['cod_tranzactie_fk'], row['tip'].capitalize(), total_suma_str, row['nr_tranzactii'])
                self.results_tree.insert("", "end", values=values, tags=(tag,))

    def _update_graph(self):
        """Desenează un grafic cu bare stivuite, RESPECTÂND filtrul de tip (D/C)."""
        self.ax.clear()

        if not self.report_data:
            self.ax.text(0.5, 0.5, 'Nicio dată de afișat conform filtrelor.', horizontalalignment='center', verticalalignment='center')
            self.canvas.draw()
            return

        df = pd.DataFrame(self.report_data)
        df['total_suma'] = pd.to_numeric(df['total_suma'])

        transaction_type_filter = self.type_var.get()
        show_credit = transaction_type_filter in ["Ambele", "Doar Credit"]
        show_debit = transaction_type_filter in ["Ambele", "Doar Debit"]

        pivot_credit = df[df['tip'] == 'credit'].pivot_table(index='perioada', columns='cod_tranzactie_fk', values='total_suma', aggfunc='sum').fillna(0)
        pivot_debit = df[df['tip'] == 'debit'].pivot_table(index='perioada', columns='cod_tranzactie_fk', values='total_suma', aggfunc='sum').fillna(0)

        all_periods = pivot_credit.index.union(pivot_debit.index)
        pivot_credit = pivot_credit.reindex(all_periods, fill_value=0)
        pivot_debit = pivot_debit.reindex(all_periods, fill_value=0)

        bar_width = 0.4 if (show_credit and show_debit and not pivot_credit.empty and not pivot_debit.empty) else 0.8

        if show_debit and not pivot_debit.empty:
            pos = 0.5 if show_credit and not pivot_credit.empty else 0
            pivot_debit.plot(kind='bar', stacked=True, ax=self.ax, color=plt.cm.Reds(np.linspace(0.4, 0.8, len(pivot_debit.columns))), width=bar_width, position=pos)

        if show_credit and not pivot_credit.empty:
            pos = -0.5 if show_debit and not pivot_debit.empty else 0
            pivot_credit.plot(kind='bar', stacked=True, ax=self.ax, color=plt.cm.Greens(np.linspace(0.4, 0.8, len(pivot_credit.columns))), width=bar_width, position=pos)

        handles, labels = self.ax.get_legend_handles_labels()
        all_tx_types_info = self.db_handler.fetch_all_dict("SELECT cod, descriere_tip FROM tipuri_tranzactii")
        self.legend_desc_map = {item['cod']: item['descriere_tip'] for item in all_tx_types_info}
        
        self.ax.legend(handles, labels, title='Coduri Tranzacție', bbox_to_anchor=(1.02, 1), loc='upper left', fontsize='small')

        self.ax.set_ylabel('Sumă Totală')
        self.ax.set_title('Analiză Tranzacții pe Perioadă')
        self.ax.grid(axis='y', linestyle='--', alpha=0.7)

        formatter = mticker.FuncFormatter(lambda x, pos: f'{x:,.0f}')
        self.ax.yaxis.set_major_formatter(formatter)

        self.ax.tick_params(axis='x', which='major', labelsize=10)
        self.ax.set_xticklabels(self.ax.get_xticklabels(), rotation=45, ha='right')

        self.fig.tight_layout(rect=[0, 0, 0.99, 1])
        self.canvas.draw()

    def _generate_pdf_file(self, output_path):
        """Generează un fișier PDF cu graficul și tabelul detaliat."""
        doc = SimpleDocTemplate(output_path, pagesize=A4, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
        styles = getSampleStyleSheet()
        story = []

        # Titlu și subtitlu
        title_text = f"Raport Analiză Tranzacții"
        story.append(Paragraph(title_text, styles['h1']))
        start_date = self.start_date_entry.get_date().strftime('%d.%m.%Y')
        end_date = self.end_date_entry.get_date().strftime('%d.%m.%Y')
        account_name = self.initial_context.get('account_name', 'N/A')
        period_text = f"Perioada: {start_date} - {end_date} | Cont: {account_name}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

        # Inserare imagine grafic
        img_buffer = io.BytesIO()
        self.fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
        img_buffer.seek(0)
        story.append(ReportlabImage(img_buffer, width=7.5*inch, height=4*inch))
        story.append(Spacer(1, 0.3*inch))

        # Inserare tabel cu date
        currency = self.initial_context.get('currency', 'RON')
        table_data = [['Perioada', 'Cod Tranz.', 'Tip', f'Sumă Totală ({currency})', 'Nr. Tranzacții']]
        
        for row in self.report_data:
            period = row['perioada']
            if isinstance(period, date):
                period = period.strftime('%d-%m-%Y')
            total_suma_str = f"{float(row['total_suma']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            table_data.append([
                period,
                row['cod_tranzactie_fk'],
                row['tip'].capitalize(),
                total_suma_str,
                row['nr_tranzactii']
            ])

        table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 2*inch, 1.5*inch])
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (3,1), (3,-1), 'RIGHT'), # Aliniere la dreapta pentru sumă
        ])
        table.setStyle(style)
        story.append(table)
        
        doc.build(story)

    def export_to_pdf(self):
        """Gestionează acțiunea de export în format PDF."""
        if not self.report_data:
            messagebox.showwarning("Export Anulat", "Nu există date de exportat.", parent=self)
            return

        account_name = self.initial_context.get('account_name', 'Raport').replace(' ', '_')
        default_filename = f"Analiza_Tranzactii_{account_name}_{date.today().strftime('%Y%m%d')}.pdf"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Fișiere PDF", "*.pdf")],
            title="Salvează Raportul PDF",
            initialfile=default_filename,
            parent=self
        )
        if not file_path:
            return
            
        try:
            self._generate_pdf_file(file_path)
            messagebox.showinfo("Succes", f"Raportul a fost salvat cu succes în:\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Export PDF", f"A apărut o eroare la salvarea fișierului PDF:\n{e}", parent=self)
            logging.error(f"Eroare la export PDF (Analiza Tranzactii): {e}", exc_info=True)

    def export_to_excel(self):
        """Gestionează acțiunea de export în format Excel."""
        if not self.report_data:
            messagebox.showwarning("Export Anulat", "Nu există date de exportat.", parent=self)
            return

        account_name = self.initial_context.get('account_name', 'Raport').replace(' ', '_')
        default_filename = f"Analiza_Tranzactii_{account_name}_{date.today().strftime('%Y%m%d')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fișiere Excel", "*.xlsx")],
            title="Salvează Raportul Excel",
            initialfile=default_filename,
            parent=self
        )
        if not file_path:
            return

        try:
            df = pd.DataFrame(self.report_data)
            df['total_suma'] = pd.to_numeric(df['total_suma'])
            if 'perioada' in df.columns and pd.api.types.is_string_dtype(df['perioada']):
                # Nu este necesară conversia dacă e deja string
                pass
            elif 'perioada' in df.columns: # Tratează cazul în care e dată
                 df['perioada'] = pd.to_datetime(df['perioada']).dt.strftime('%d.%m.%Y')


            df.rename(columns={
                'perioada': 'Perioada', 'cod_tranzactie_fk': 'Cod Tranzacție',
                'tip': 'Tip', 'total_suma': 'Sumă Totală', 'nr_tranzactii': 'Nr. Tranzacții'
            }, inplace=True)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                sheet_name = "Analiza Tranzactii"
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                ws = writer.sheets[sheet_name]
                
                title_font = ExcelFont(bold=True, size=14, name='Calibri')
                header_font = ExcelFont(bold=True, color="FFFFFF", name='Calibri', size=11)
                header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                start_date_str = self.start_date_entry.get_date().strftime('%d.%m.%Y')
                end_date_str = self.end_date_entry.get_date().strftime('%d.%m.%Y')
                title_text = f"Raport Analiză Tranzacții: {self.initial_context.get('account_name', 'N/A')} ({start_date_str} - {end_date_str})"
                title_cell = ws['A1']
                title_cell.value = title_text; title_cell.font = title_font
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                title_cell.alignment = Alignment(horizontal='center')

                for cell in ws[2]: cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
                
                ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 15

                for row_cells in ws.iter_rows(min_row=3, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                    for cell in row_cells:
                        cell.border = thin_border
                        if cell.column == 4: cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal='right')
                
                img_buffer = io.BytesIO()
                self.fig.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
                img = OpenpyxlImage(img_buffer)
                img.anchor = 'G2'
                ws.add_image(img)

            messagebox.showinfo("Succes", f"Raportul a fost salvat cu succes în:\n{file_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Export Excel", f"A apărut o eroare la salvarea fișierului:\n{e}", parent=self)
            logging.error(f"Eroare la export Excel (Analiza Tranzactii): {e}", exc_info=True)
            
    def _on_send_email(self):
        """Gestionează acțiunea de trimitere a raportului pe email."""
        if not self.report_data:
            messagebox.showwarning("Acțiune Anulată", "Vă rugăm mai întâi generați un raport.", parent=self)
            return

        recipient = simpledialog.askstring("Adresă Destinatar", "Introduceți adresa de email a destinatarului:", parent=self)
        if not recipient: return

        try:
            account_name = self.initial_context.get('account_name', 'N/A')
            start_date_str = self.start_date_entry.get_date().strftime('%d.%m.%Y')
            end_date_str = self.end_date_entry.get_date().strftime('%d.%m.%Y')
            currency = self.initial_context.get('currency', 'RON')
            
            total_credit = sum(float(r['total_suma']) for r in self.report_data if r['tip'] == 'credit')
            total_debit = sum(float(r['total_suma']) for r in self.report_data if r['tip'] == 'debit')
            f_credit = f"{total_credit:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            f_debit = f"{total_debit:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            subject = f"Analiză Tranzacții: {account_name} ({start_date_str} - {end_date_str})"
            body = f"""Bună ziua,\n\nAtașat găsiți raportul de analiză detaliată a tranzacțiilor generat din {APP_NAME}.\n
--- REZUMAT RAPORT ---
Cont Analizat: {account_name}
Perioada: {start_date_str} - {end_date_str}
Tip Tranzacții: {self.type_var.get()}
Granularitate: {self.granularity_var.get()}
--------------------
Total Credit pe categorii: {f_credit} {currency}
Total Debit pe categorii: {f_debit} {currency}
--------------------\n
O zi bună,
Email generat automat de {APP_NAME} v{APP_VERSION}
"""

            temp_dir = tempfile.gettempdir()
            temp_pdf_path = os.path.join(temp_dir, f"Analiza_Tranzactii_{account_name.replace(' ', '_')}_{date.today().strftime('%Y%m%d')}.pdf")
        
            self._generate_pdf_file(temp_pdf_path)
            success, message = send_report_email(self.smtp_config, recipient, subject, body, temp_pdf_path)

            if success: messagebox.showinfo("Succes", message, parent=self)
            else: messagebox.showerror("Eroare Trimitere", message, parent=self)
        except Exception as e:
            messagebox.showerror("Eroare Critică", f"A apărut o eroare neașteptată:\n{e}", parent=self)
            logging.error(f"Eroare la trimitere email (Analiza Tranzactii): {e}", exc_info=True)
        finally:
            if os.path.exists(temp_pdf_path):
                try: os.remove(temp_pdf_path)
                except OSError as e: logging.warning(f"Nu s-a putut șterge fișierul PDF temporar: {temp_pdf_path}. Eroare: {e}")